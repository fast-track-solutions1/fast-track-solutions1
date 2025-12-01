# api/import_utils.py - LOGIQUE D'IMPORTATION GÉNÉRIQUE

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import logging
from django.apps import apps
from django.db import transaction
from django.core.exceptions import ValidationError
from datetime import datetime
import json

logger = logging.getLogger(__name__)

# ============================================================================
# CONFIGURATION DES MODÈLES IMPORTABLES
# ============================================================================

IMPORTABLE_MODELS = {
    'societe': {
        'app': 'api',
        'model': 'Societe',
        'unique_field': 'nom',
        'exclude_fields': ['id', 'date_creation'],
    },
    'departement': {
        'app': 'api',
        'model': 'Departement',
        'unique_field': 'numero',
        'exclude_fields': ['id', 'date_creation'],
    },
    'circuit': {
        'app': 'api',
        'model': 'Circuit',
        'unique_field': None,
        'exclude_fields': ['id', 'date_creation'],
    },
    'service': {
        'app': 'api',
        'model': 'Service',
        'unique_field': 'nom',
        'exclude_fields': ['id', 'date_creation'],
    },
    'grade': {
        'app': 'api',
        'model': 'Grade',
        'unique_field': 'nom',
        'exclude_fields': ['id', 'date_creation'],
    },
    'creneau_travail': {
        'app': 'api',
        'model': 'CreneauTravail',
        'unique_field': 'nom',
        'exclude_fields': ['id', 'date_creation'],
    },
    'type_acces': {
        'app': 'api',
        'model': 'TypeAcces',
        'unique_field': 'nom',
        'exclude_fields': ['id'],
    },
    'outil_travail': {
        'app': 'api',
        'model': 'OutilTravail',
        'unique_field': 'nom',
        'exclude_fields': ['id'],
    },
    'equipement': {
        'app': 'api',
        'model': 'Equipement',
        'unique_field': None,
        'exclude_fields': ['id', 'date_creation'],
    },
    'salarie': {
        'app': 'api',
        'model': 'Salarie',
        'unique_field': 'matricule',
        'exclude_fields': ['id', 'date_creation', 'date_modification'],
    },
}


class GenericImporter:
    """Classe générique pour importer n'importe quel modèle Django depuis Excel"""

    def __init__(self, model_name: str):
        """
        Initialise l'importeur
        
        Args:
            model_name: clé du modèle à importer (ex: 'salarie', 'departement')
        """
        if model_name not in IMPORTABLE_MODELS:
            raise ValueError(f"Modèle '{model_name}' non importable. Disponibles: {list(IMPORTABLE_MODELS.keys())}")
        
        self.model_name = model_name
        self.config = IMPORTABLE_MODELS[model_name]
        self.Model = apps.get_model(self.config['app'], self.config['model'])
        self.results = {
            'inserted': 0,
            'updated': 0,
            'errors': [],
            'warnings': []
        }

    def generate_template(self) -> bytes:
        """
        Génère un fichier Excel template basé sur les champs du modèle
        
        Returns:
            BytesIO: Fichier Excel en bytes
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Données"

            # Récupérer tous les champs
            fields = self._get_importable_fields()
            
            # Style pour l'en-tête
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Ajouter les en-têtes
            for col_num, field_name in enumerate(fields, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = field_name
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Ajouter 5 lignes vides d'exemple
            for row in range(2, 7):
                for col in range(1, len(fields) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border

            # Ajuster la largeur des colonnes
            for col_num, field_name in enumerate(fields, 1):
                ws.column_dimensions[chr(64 + col_num)].width = len(field_name) + 5

            # Figer l'en-tête
            ws.freeze_panes = "A2"

            # Sauvegarder dans BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        except Exception as e:
            logger.error(f"Erreur lors de la génération du template: {str(e)}")
            raise

    def import_from_excel(self, file) -> dict:
        """
        Importe les données depuis un fichier Excel
        
        Args:
            file: Fichier Excel uploadé
            
        Returns:
            dict: Résultat de l'import {inserted, updated, errors, warnings}
        """
        try:
            # Lire le fichier Excel
            df = pd.read_excel(file)
            
            if df.empty:
                raise ValueError("Le fichier Excel est vide")

            # Nettoyer les données
            df = df.fillna('')  # Remplacer NaN par chaîne vide
            df.columns = [col.strip().lower().replace(' ', '_') for col in df.columns]  # Normaliser les colonnes

            logger.info(f"Import de {len(df)} lignes pour {self.model_name}")

            # Importer chaque ligne
            with transaction.atomic():
                for idx, row in df.iterrows():
                    try:
                        self._import_row(row, idx + 2)  # +2 car idx commence à 0 et ligne 1 est l'en-tête
                    except Exception as e:
                        self.results['errors'].append({
                            'row': idx + 2,
                            'error': str(e)
                        })
                        logger.error(f"Erreur ligne {idx + 2}: {str(e)}")

            return self.results

        except Exception as e:
            logger.error(f"Erreur lors de l'import: {str(e)}")
            self.results['errors'].append({'row': 0, 'error': f"Erreur générale: {str(e)}"})
            return self.results

    def _import_row(self, row: pd.Series, row_num: int):
        """
        Importe une ligne unique
        
        Args:
            row: Série pandas représentant la ligne
            row_num: Numéro de la ligne (pour les erreurs)
        """
        # Préparer les données
        data = {}
        for key, value in row.items():
            if pd.isna(value) or value == '':
                continue
            
            # Valider et convertir le champ
            data[key] = self._convert_field_value(key, value)

        if not data:
            self.results['warnings'].append({'row': row_num, 'warning': 'Ligne vide'})
            return

        # Déterminer clé unique
        unique_field = self.config.get('unique_field')
        update_or_create_kwargs = {}
        
        if unique_field and unique_field in data:
            update_or_create_kwargs[unique_field] = data[unique_field]
        elif 'id' in data and data['id']:
            update_or_create_kwargs['id'] = data['id']
        else:
            # Aucune clé unique, créer directement
            obj = self.Model.objects.create(**data)
            self.results['inserted'] += 1
            return

        # Update or create
        defaults = {k: v for k, v in data.items() if k not in update_or_create_kwargs}
        obj, created = self.Model.objects.update_or_create(
            **update_or_create_kwargs,
            defaults=defaults
        )

        if created:
            self.results['inserted'] += 1
        else:
            self.results['updated'] += 1

    def _convert_field_value(self, field_name: str, value):
        """
        Convertit la valeur en type approprié
        
        Args:
            field_name: Nom du champ
            value: Valeur brute du fichier Excel
            
        Returns:
            Valeur convertie appropriée
        """
        try:
            # Récupérer le champ du modèle
            field = self.Model._meta.get_field(field_name)
            field_type = field.get_internal_type()

            if field_type == 'ForeignKey':
                # Résoudre la relation
                related_model = field.related_model
                try:
                    return related_model.objects.get(id=int(value))
                except (ValueError, related_model.DoesNotExist):
                    raise ValueError(f"Impossible de trouver {related_model.__name__} avec id={value}")

            elif field_type == 'ManyToManyField':
                # Gérer les M2M (pas supporté pour l'instant)
                raise ValueError("ManyToMany non supporté pour l'import")

            elif field_type in ['DateField', 'DateTimeField']:
                if isinstance(value, str):
                    # Essayer différents formats
                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                        try:
                            return datetime.strptime(value, fmt).date() if field_type == 'DateField' else datetime.strptime(value, fmt)
                        except ValueError:
                            continue
                    raise ValueError(f"Format de date invalide: {value}")
                return value

            elif field_type == 'BooleanField':
                if isinstance(value, str):
                    return value.lower() in ['true', '1', 'yes', 'oui', 'vrai']
                return bool(value)

            elif field_type == 'IntegerField':
                return int(value)

            elif field_type == 'FloatField':
                return float(value)

            elif field_type == 'EmailField':
                value_str = str(value).strip()
                if '@' not in value_str:
                    raise ValueError(f"Email invalide: {value}")
                return value_str

            else:
                return str(value).strip()

        except Exception as e:
            raise ValueError(f"Erreur conversion champ '{field_name}': {str(e)}")

    def _get_importable_fields(self) -> list:
        """
        Récupère la liste des champs importables
        
        Returns:
            list: Liste des noms de champs importables
        """
        exclude = self.config.get('exclude_fields', [])
        fields = [
            f.name for f in self.Model._meta.get_fields()
            if f.name not in exclude and not f.many_to_one and not f.many_to_many
        ]
        return fields

    def get_model_structure(self) -> dict:
        """
        Retourne la structure du modèle (champs, types, etc.)
        
        Returns:
            dict: Description de la structure
        """
        structure = {
            'model': self.model_name,
            'fields': []
        }

        for field in self.Model._meta.get_fields():
            if field.name in self.config.get('exclude_fields', []):
                continue

            field_info = {
                'name': field.name,
                'type': field.get_internal_type(),
                'required': not field.null and not field.blank,
                'help_text': field.help_text if hasattr(field, 'help_text') else '',
            }

            if hasattr(field, 'choices') and field.choices:
                field_info['choices'] = [{'value': v, 'label': l} for v, l in field.choices]

            if field.get_internal_type() == 'ForeignKey':
                field_info['related_model'] = field.related_model.__name__

            structure['fields'].append(field_info)

        return structure


def get_importable_models() -> dict:
    """Retourne la liste de tous les modèles importables"""
    return {
        key: {
            'name': config['model'],
            'app': config['app']
        }
        for key, config in IMPORTABLE_MODELS.items()
    }
