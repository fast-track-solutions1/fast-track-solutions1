import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import getcolumnletter
from io import BytesIO
import logging
from django.apps import apps
from django.db import transaction
from datetime import datetime

logger = logging.getLogger(__name__)

IMPORTABLE_MODELS = {
    'societe': {'app': 'api', 'model': 'Societe', 'name': 'Société', 'unique_field': 'nom', 'exclude_fields': ['id', 'date_creation']},
    'departement': {'app': 'api', 'model': 'Departement', 'name': 'Département', 'unique_field': 'numero', 'exclude_fields': ['id', 'date_creation']},
    'circuit': {'app': 'api', 'model': 'Circuit', 'name': 'Circuit', 'unique_field': None, 'exclude_fields': ['id', 'date_creation']},
    'service': {'app': 'api', 'model': 'Service', 'name': 'Service', 'unique_field': 'nom', 'exclude_fields': ['id', 'date_creation']},
    'grade': {'app': 'api', 'model': 'Grade', 'name': 'Grade', 'unique_field': 'nom', 'exclude_fields': ['id', 'date_creation']},
    'creneautravail': {'app': 'api', 'model': 'CreneauTravail', 'name': 'Créneau de Travail', 'unique_field': 'nom', 'exclude_fields': ['id', 'date_creation']},
    'typeacces': {'app': 'api', 'model': 'TypeAcces', 'name': "Type d'Accès", 'unique_field': 'nom', 'exclude_fields': ['id']},
    'outiltravail': {'app': 'api', 'model': 'OutilTravail', 'name': 'Outil de Travail', 'unique_field': 'nom', 'exclude_fields': ['id']},
    'equipement': {'app': 'api', 'model': 'Equipement', 'name': 'Équipement', 'unique_field': None, 'exclude_fields': ['id', 'date_creation']},
    'salarie': {'app': 'api', 'model': 'Salarie', 'name': 'Salarié', 'unique_field': 'matricule', 'exclude_fields': ['id', 'date_creation', 'date_modification']},
    'accesapplication': {'app': 'api', 'model': 'AccesApplication', 'name': 'Accès Application', 'unique_field': None, 'exclude_fields': ['id', 'date_creation']},
    'equipementinstance': {'app': 'api', 'model': 'EquipementInstance', 'name': 'Équipement Instance', 'unique_field': 'numero_serie', 'exclude_fields': ['id', 'date_creation']},
    'horairesalarie': {'app': 'api', 'model': 'HoraireSalarie', 'name': 'Horaire Salarié', 'unique_field': None, 'exclude_fields': ['id', 'date_creation']},
}

def get_importable_models():
    return IMPORTABLE_MODELS

class GenericImporter:
    def __init__(self, model_name: str):
        if model_name not in IMPORTABLE_MODELS:
            raise ValueError(f"Modèle {model_name} non importable. Disponibles: {list(IMPORTABLE_MODELS.keys())}")
        self.model_name = model_name
        self.config = IMPORTABLE_MODELS[model_name]
        self.Model = apps.get_model(self.config['app'], self.config['model'])
        self.results = {'inserted': 0, 'updated': 0, 'errors': [], 'warnings': []}
    
    def get_model_structure(self) -> dict:
        try:
            fields = self.get_importable_fields()
            structure = {'fields': [], 'unique_field': self.config.get('unique_field'), 'exclude_fields': self.config.get('exclude_fields', [])}
            for field_name in fields:
                field = self.Model._meta.get_field(field_name)
                field_info = {'name': field_name, 'type': field.get_internal_type(), 'required': not field.null and not field.blank}
                if field.get_internal_type() == 'ForeignKey':
                    related_model = field.related_model
                    field_info['is_foreign_key'] = True
                    field_info['related_model'] = related_model.__name__
                    try:
                        values = list(related_model.objects.values_list('nom', flat=True))
                        field_info['possible_values'] = values
                    except:
                        field_info['possible_values'] = []
                else:
                    field_info['is_foreign_key'] = False
                structure['fields'].append(field_info)
            return structure
        except Exception as e:
            logger.error(f"Erreur structure: {str(e)}")
            raise
    
    def get_importable_fields(self) -> list:
        exclude = self.config.get('exclude_fields', [])
        return [f.name for f in self.Model._meta.get_fields() if f.name not in exclude and not f.many_to_many and not (f.auto_created and not f.concrete)]
    
    def generate_template(self) -> bytes:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Données"
            fields = self.get_importable_fields()
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for col_num, field_name in enumerate(fields, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = field_name
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for row in range(2, 12):
                for col in range(1, len(fields) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
            for col_num, field_name in enumerate(fields, 1):
                ws.column_dimensions[getcolumnletter(col_num)].width = max(15, len(field_name) + 5)
            ws.freeze_panes = "A2"
            for col_num, field_name in enumerate(fields, 1):
                try:
                    field = self.Model._meta.get_field(field_name)
                    if field.get_internal_type() == 'ForeignKey':
                        related_model = field.related_model
                        values = list(related_model.objects.values_list('nom', flat=True).order_by('nom'))
                        if values:
                            dv = DataValidation(type="list", formula1=f'"{",".join(map(str, values))}"', allow_blank=True)
                            dv.error = f'Sélectionnez une valeur valide pour {field_name}'
                            dv.errorTitle = f'Valeur invalide: {field_name}'
                            ws.add_data_validation(dv)
                            col_letter = getcolumnletter(col_num)
                            for row in range(2, 1001):
                                dv.add(f'{col_letter}{row}')
                except Exception as e:
                    logger.warning(f"Impossible liste déroulante {field_name}: {str(e)}")
            instructions_sheet = wb.create_sheet("Instructions")
            instructions_sheet['A1'] = "INSTRUCTIONS D'IMPORT"
            instructions_sheet['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            instructions_sheet['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            row = 3
            instructions_sheet[f'A{row}'] = "Colonnes obligatoires"
            instructions_sheet[f'A{row}'].font = Font(bold=True, size=11)
            row += 1
            unique_field = self.config.get('unique_field')
            for field_name in fields:
                try:
                    field = self.Model._meta.get_field(field_name)
                    is_required = not field.null and not field.blank
                    is_unique = field_name == unique_field
                    if is_required or is_unique:
                        marker = "*" if is_required else ""
                        instructions_sheet[f'A{row}'] = f"{marker} {field_name}"
                        if is_unique:
                            instructions_sheet[f'B{row}'] = "Clé unique"
                        row += 1
                except:
                    pass
            instructions_sheet.column_dimensions['A'].width = 30
            instructions_sheet.column_dimensions['B'].width = 80
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
        except Exception as e:
            logger.error(f"Erreur template: {str(e)}")
            raise
    
    def import_from_excel(self, file) -> dict:
        try:
            df = pd.read_excel(file)
            if df.empty:
                raise ValueError("Le fichier Excel est vide")
            df = df.fillna('')
            df.columns = [col.strip().lower().replace(' ', '_') for col in df.columns]
            logger.info(f"Import de {len(df)} lignes pour {self.model_name}")
            with transaction.atomic():
                for idx, row in df.iterrows():
                    try:
                        self.import_row(row, idx + 2)
                    except Exception as e:
                        self.results['errors'].append({'row': idx + 2, 'error': str(e)})
                        logger.error(f"Erreur ligne {idx + 2}: {str(e)}")
            return self.results
        except Exception as e:
            logger.error(f"Erreur import: {str(e)}")
            self.results['errors'].append({'row': 0, 'error': f"Erreur générale: {str(e)}"})
            return self.results
    
    def import_row(self, row: pd.Series, row_num: int):
        data = {}
        for key, value in row.items():
            if pd.isna(value) or value == '':
                continue
            data[key] = self.convert_field_value(key, value)
        if not data:
            self.results['warnings'].append({'row': row_num, 'warning': 'Ligne vide'})
            return
        unique_field = self.config.get('unique_field')
        update_or_create_kwargs = {}
        if unique_field and unique_field in data:
            update_or_create_kwargs[unique_field] = data[unique_field]
        elif 'id' in data and data['id']:
            update_or_create_kwargs['id'] = data['id']
        else:
            obj = self.Model.objects.create(**data)
            self.results['inserted'] += 1
            return
        defaults = {k: v for k, v in data.items() if k not in update_or_create_kwargs}
        obj, created = self.Model.objects.update_or_create(update_or_create_kwargs, defaults=defaults)
        if created:
            self.results['inserted'] += 1
        else:
            self.results['updated'] += 1
    
    def convert_field_value(self, field_name: str, value):
        try:
            field = self.Model._meta.get_field(field_name)
            field_type = field.get_internal_type()
            if field_type == 'ForeignKey':
                related_model = field.related_model
                try:
                    return related_model.objects.get(nom=str(value).strip())
                except related_model.DoesNotExist:
                    try:
                        return related_model.objects.get(id=int(value))
                    except (ValueError, related_model.DoesNotExist):
                        raise ValueError(f"Impossible de trouver {related_model.__name__} avec: {value}")
            elif field_type in ['DateField', 'DateTimeField']:
                if isinstance(value, str):
                    try:
                        return datetime.strptime(value, '%Y-%m-%d').date()
                    except ValueError:
                        try:
                            return datetime.strptime(value, '%d/%m/%Y').date()
                        except ValueError:
                            raise ValueError(f"Format date invalide: {value}")
                return value
            elif field_type == 'BooleanField':
                if isinstance(value, str):
                    return value.lower() in ['true', '1', 'yes', 'oui']
                return bool(value)
            elif field_type in ['IntegerField', 'AutoField', 'BigIntegerField', 'SmallIntegerField']:
                return int(value)
            elif field_type in ['DecimalField', 'FloatField']:
                return float(value)
            return str(value).strip()
        except Exception as e:
            raise ValueError(f"Erreur conversion {field_name}: {str(e)}")
