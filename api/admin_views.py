from django.shortcuts import render, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
import requests
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .forms import BulkImportForm

@staff_member_required
def admin_import_page(request):
    """
    Page d'import en masse dans Django Admin
    Accessible à /admin/import/
    """
    result = None
    error = None

    if request.method == 'POST':
        form = BulkImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                api_name = form.cleaned_data['api_name']
                file = request.FILES['file']
                skip_errors = form.cleaned_data.get('skip_errors', False)

                # 📤 Préparer le fichier pour l'upload
                files = {'file': file}
                params = {
                    'api_name': api_name,
                    'skip_errors': 'true' if skip_errors else 'false'
                }

                # 🔑 Récupérer le token JWT de l'utilisateur
                from rest_framework_simplejwt.tokens import RefreshToken
                refresh = RefreshToken.for_user(request.user)
                access_token = str(refresh.access_token)

                # 📡 Appeler l'endpoint d'import
                headers = {
                    'Authorization': f'Bearer {access_token}'
                }

                # ✅ CORRIGER L'URL AVEC LE PARAMÈTRE EN URL PATH
                response = requests.post(
                    f'http://localhost:8000/api/import/upload/{api_name}/',
                    files=files,
                    headers=headers,
                    timeout=30
                )

                if response.status_code == 200:
                    result = response.json()

                    # 📊 Afficher un message de succès
                    if result.get('statut') == 'succes':
                        messages.success(
                            request,
                            f"✅ Import réussi ! {result.get('lignes_succes')} ligne(s) importée(s)."
                        )
                    else:
                        messages.warning(
                            request,
                            f"⚠️ Import partiel : {result.get('lignes_succes')} succès, {result.get('lignes_erreur')} erreur(s)."
                        )
                else:
                    error_data = response.json()
                    error = f"❌ Erreur: {error_data.get('error', 'Erreur inconnue')}"
                    messages.error(request, error)

            except requests.exceptions.ConnectionError:
                error = "❌ Impossible de se connecter à l'API. Vérifiez que le serveur Django est lancé."
                messages.error(request, error)
            except Exception as e:
                error = f"❌ Erreur lors de l'import: {str(e)}"
                messages.error(request, error)
    else:
        form = BulkImportForm()

    context = {
        'form': form,
        'result': result,
        'error': error,
        'title': '📥 Importation en Masse',
    }

    return render(request, 'admin/import_page.html', context)


@staff_member_required
def admin_download_template(request):
    """
    Télécharge un template Excel basé sur le modèle sélectionné
    """
    api_name = request.GET.get('api_name', '')

    if not api_name:
        return HttpResponse('❌ Modèle non spécifié', status=400)

    # 🔑 Récupérer le token JWT
    from rest_framework_simplejwt.tokens import RefreshToken
    refresh = RefreshToken.for_user(request.user)
    access_token = str(refresh.access_token)

    # 📡 Récupérer la structure du modèle
    headers = {'Authorization': f'Bearer {access_token}'}

    try:
        # ✅ CORRIGER L'URL AVEC LE PARAMÈTRE EN URL PATH
        response = requests.get(
            f'http://localhost:8000/api/import/structure/{api_name}/',
            headers=headers,
            timeout=10
        )

        if response.status_code != 200:
            return HttpResponse('❌ Impossible de récupérer la structure', status=400)

        structure = response.json()
        fields = structure.get('fields', [])

        # 📊 Créer le fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Données"

        # En-têtes
        header_fill = PatternFill(start_color="217346", end_color="217346", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = field
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-ajuster les largeurs
        for col_idx, field in enumerate(fields, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = len(field) + 2

        # Exemple de ligne vide
        for col_idx in range(1, len(fields) + 1):
            ws.cell(row=2, column=col_idx).value = ""

        # Sauvegarder
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="template_{api_name}.xlsx"'
        wb.save(response)

        return response

    except Exception as e:
        return HttpResponse(f'❌ Erreur: {str(e)}', status=500)
