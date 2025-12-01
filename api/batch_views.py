from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import csv
import json
from io import StringIO, BytesIO
from django.apps import apps
from django.db.models import ForeignKey, ManyToManyField
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# ============================================================================
# CONFIGURATION - MODÈLES ET CHAMPS À IGNORER
# ============================================================================

IGNORED_FIELDS = {
    'id', 'pk', 'date_creation', 'date_modification', 'date_dernière_maj',
    'password', 'mot_de_passe', 'created_at', 'updated_at', 'date_update'
}

# Clés uniques par modèle pour dédoublonner à l'import
UNIQUE_KEYS = {
    'Societe': 'nom',
    'Departement': ('societe', 'numero'),
    'Circuit': ('departement', 'nom'),
    'Service': ('societe', 'nom'),
    'Grade': ('societe', 'nom'),
    'TypeAcces': 'nom',
    'OutilTravail': 'nom',
    'CreneauTravail': ('societe', 'nom'),
    'Equipement': ('nom', 'type_equipement'),
    'Salarie': ('societe', 'matricule'),
    'TypeApplicationAcces': 'nom',
    'EquipementInstance': 'numero_serie',
    'AccesApplication': ('salarie', 'application'),
    'AccesSalarie': ('salarie', 'type_acces'),
    'DemandeConge': None,
    'DemandeAcompte': None,
    'DemandeSortie': None,
    'TravauxExceptionnels': None,
    'DocumentSalarie': None,
    'HistoriqueSalarie': None,
    'HoraireSalarie': None,
    'FichePoste': ('service', 'titre'),
    'AmeliorationProposee': None,
    'OutilFichePoste': ('fiche_poste', 'outil_travail'),
}

# ============================================================================
# UTILITAIRES
# ============================================================================

def get_model_fields(model, exclude_auto=True):
    """
    Récupère les champs exportables du modèle
    - Exclut les champs auto-generated (id, date_creation, etc.)
    - Exclut les ForeignKey et ManyToMany (traités séparément)
    """
    fields = []
    for field in model._meta.fields:
        # Ignorer les champs auto
        if exclude_auto and field.name in IGNORED_FIELDS:
            continue
        # Ignorer auto_now et auto_now_add
        if hasattr(field, 'auto_now') and (field.auto_now or field.auto_now_add):
            continue
        # Ignorer les relations (ForeignKey, ManyToMany)
        if isinstance(field, ForeignKey):
            # Garder seulement le champ _id pour les FK
            fields.append(field.name + '_id')
        elif not isinstance(field, ManyToManyField):
            fields.append(field.name)
    return fields

def get_unique_key_for_model(model_name):
    """Retourne la(les) clé(s) unique(s) pour un modèle"""
    return UNIQUE_KEYS.get(model_name, None)

def parse_field_value(field, value, model):
    """
    Parse la valeur d'un champ selon son type
    Gère les ForeignKey, bool, datetime, etc.
    """
    if not value or value == '':
        return None
    
    try:
        # Récupérer le champ Django
        field_obj = model._meta.get_field(field)
        
        # ForeignKey - récupérer l'ID
        if isinstance(field_obj, ForeignKey):
            try:
                return int(value)
            except (ValueError, TypeError):
                return None
        
        # Boolean
        if field_obj.get_internal_type() == 'BooleanField':
            return str(value).lower() in ('true', '1', 'oui', 'yes', 'o')
        
        # Date
        if field_obj.get_internal_type() == 'DateField':
            if isinstance(value, str):
                try:
                    return datetime.strptime(value, '%Y-%m-%d').date()
                except ValueError:
                    try:
                        return datetime.strptime(value, '%d/%m/%Y').date()
                    except ValueError:
                        return None
            return value
        
        # DateTime
        if field_obj.get_internal_type() == 'DateTimeField':
            if isinstance(value, str):
                try:
                    return datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    return None
            return value
        
        # Decimal
        if field_obj.get_internal_type() == 'DecimalField':
            try:
                return float(value)
            except (ValueError, TypeError):
                return None
        
        # Integer
        if field_obj.get_internal_type() in ['IntegerField', 'AutoField']:
            try:
                return int(value)
            except (ValueError, TypeError):
                return None
        
        # String par défaut
        return str(value)
    
    except Exception as e:
        return value

def validate_row_data(model, row_data, row_num):
    """
    Valide les données d'une ligne
    Retourne (is_valid, cleaned_data, errors)
    """
    errors = []
    cleaned_data = {}
    
    for field_name, value in row_data.items():
        if not value or value == '':
            continue
        
        try:
            field_obj = model._meta.get_field(field_name)
            parsed_value = parse_field_value(field_name, value, model)
            cleaned_data[field_name] = parsed_value
        except Exception as e:
            errors.append(f"Champ '{field_name}': {str(e)}")
    
    return (len(errors) == 0, cleaned_data, errors)

# ============================================================================
# API ENDPOINTS - VUES DJANGO CLASSIQUES (pas DRF)
# ============================================================================

@csrf_exempt
@require_http_methods(["GET"])
def batch_template(request, model_name):
    """
    Télécharger un template CSV/Excel pour un modèle
    GET /api/batch/template/<model_name>/?format=csv|xlsx
    """
    format_type = request.GET.get('format', 'csv').lower()
    
    # Valider le modèle
    try:
        Model = apps.get_model('api', model_name)
    except LookupError:
        return HttpResponse(
            json.dumps({'error': f'Modèle {model_name} non trouvé'}),
            status=404,
            content_type='application/json'
        )
    
    # Récupérer les champs exportables
    fields = get_model_fields(Model, exclude_auto=True)
    
    if format_type == 'xlsx':
        return _generate_excel_template(model_name, fields)
    else:
        return _generate_csv_template(model_name, fields)

def _generate_csv_template(model_name, fields):
    """Génère un template CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{model_name}_template.csv"'
    
    writer = csv.writer(response)
    writer.writerow(fields)
    
    return response

def _generate_excel_template(model_name, fields):
    """Génère un template Excel avec mise en forme"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = model_name
    
    # En-tête avec style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col_num, field_name in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = field_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = 20
    
    # Ajouter 5 lignes vides pour l'exemple
    for row in range(2, 7):
        for col in range(1, len(fields) + 1):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="left")
    
    # Retourner le fichier
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{model_name}_template.xlsx"'
    
    return response

@csrf_exempt
@require_http_methods(["POST"])
def batch_import(request, model_name):
    """
    Importer des données pour un modèle
    POST /api/batch/import/<model_name>/
    """
    
    # Valider le modèle
    try:
        Model = apps.get_model('api', model_name)
    except LookupError:
        return HttpResponse(
            json.dumps({'error': f'Modèle {model_name} non trouvé'}),
            status=404,
            content_type='application/json'
        )
    
    # Vérifier le fichier
    if 'file' not in request.FILES:
        return HttpResponse(
            json.dumps({'error': 'Aucun fichier fourni'}),
            status=400,
            content_type='application/json'
        )
    
    file = request.FILES['file']
    dry_run = request.POST.get('dry_run', 'false').lower() == 'true'
    
    # Parser le fichier
    if file.name.endswith('.xlsx'):
        data = _parse_excel_file(file)
    elif file.name.endswith('.csv'):
        data = _parse_csv_file(file)
    else:
        return HttpResponse(
            json.dumps({'error': 'Format non supporté. Utilisez CSV ou XLSX.'}),
            status=400,
            content_type='application/json'
        )
    
    if not data or 'error' in data:
        return HttpResponse(
            json.dumps(data or {'error': 'Erreur de parsing du fichier'}),
            status=400,
            content_type='application/json'
        )
    
    # Importer les données
    results = _process_import(Model, data, dry_run)
    
    return HttpResponse(
        json.dumps(results),
        status=200,
        content_type='application/json'
    )

def _parse_csv_file(file):
    """Parse un fichier CSV"""
    try:
        file_data = file.read().decode('utf-8')
        csv_reader = csv.DictReader(StringIO(file_data))
        data = list(csv_reader)
        return data
    except Exception as e:
        return {'error': f'Erreur CSV: {str(e)}'}

def _parse_excel_file(file):
    """Parse un fichier Excel"""
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        # Récupérer l'en-tête
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Récupérer les données
        data = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            row_dict = {}
            for col_num, cell in enumerate(row):
                header = headers[col_num]
                row_dict[header] = cell.value
            # Ignorer les lignes vides
            if any(row_dict.values()):
                data.append(row_dict)
        
        return data
    except Exception as e:
        return {'error': f'Erreur XLSX: {str(e)}'}

def _process_import(model, rows_data, dry_run=False):
    """
    Traite l'import ligne par ligne
    """
    results = []
    stats = {'created': 0, 'updated': 0, 'errors': 0, 'total': len(rows_data)}
    
    unique_key = get_unique_key_for_model(model.__name__)
    
    for row_num, row_data in enumerate(rows_data, start=2):
        
        # Valider les données
        is_valid, cleaned_data, validation_errors = validate_row_data(model, row_data, row_num)
        
        if not is_valid:
            results.append({
                'row': row_num,
                'status': 'error',
                'errors': validation_errors
            })
            stats['errors'] += 1
            continue
        
        try:
            # Cas 1: Pas de clé unique (toujours créer)
            if unique_key is None:
                if not dry_run:
                    obj = model.objects.create(**cleaned_data)
                results.append({
                    'row': row_num,
                    'status': 'created',
                    'id': obj.id if not dry_run else 'N/A (dry_run)',
                    'message': 'OK'
                })
                stats['created'] += 1
            
            # Cas 2: Une seule clé unique
            elif isinstance(unique_key, str):
                key_value = cleaned_data.get(unique_key)
                if key_value is None:
                    results.append({
                        'row': row_num,
                        'status': 'error',
                        'errors': [f"Champ clé '{unique_key}' manquant ou vide"]
                    })
                    stats['errors'] += 1
                    continue
                
                if not dry_run:
                    obj, created = model.objects.update_or_create(
                        **{unique_key: key_value},
                        defaults=cleaned_data
                    )
                else:
                    exists = model.objects.filter(**{unique_key: key_value}).exists()
                    created = not exists
                
                results.append({
                    'row': row_num,
                    'status': 'created' if created else 'updated',
                    'id': obj.id if not dry_run else 'N/A (dry_run)',
                    'message': 'OK'
                })
                stats['created' if created else 'updated'] += 1
            
            # Cas 3: Clé unique composite (tuple)
            else:
                lookup_dict = {}
                for key_field in unique_key:
                    key_value = cleaned_data.get(key_field)
                    if key_value is None:
                        results.append({
                            'row': row_num,
                            'status': 'error',
                            'errors': [f"Champ clé '{key_field}' manquant ou vide"]
                        })
                        stats['errors'] += 1
                        raise Exception('Missing key')
                    lookup_dict[key_field] = key_value
                
                if not dry_run:
                    obj, created = model.objects.update_or_create(
                        **lookup_dict,
                        defaults=cleaned_data
                    )
                else:
                    exists = model.objects.filter(**lookup_dict).exists()
                    created = not exists
                
                results.append({
                    'row': row_num,
                    'status': 'created' if created else 'updated',
                    'id': obj.id if not dry_run else 'N/A (dry_run)',
                    'message': 'OK'
                })
                stats['created' if created else 'updated'] += 1
        
        except Exception as e:
            results.append({
                'row': row_num,
                'status': 'error',
                'errors': [str(e)]
            })
            stats['errors'] += 1
    
    return {
        'success': stats['errors'] == 0,
        'total_rows': stats['total'],
        'created': stats['created'],
        'updated': stats['updated'],
        'errors': stats['errors'],
        'dry_run': dry_run,
        'results': results
    }

@csrf_exempt
@require_http_methods(["GET"])
def batch_export(request, model_name):
    """
    Exporter tous les enregistrements d'un modèle
    GET /api/batch/export/<model_name>/?format=csv|xlsx
    """
    format_type = request.GET.get('format', 'csv').lower()
    
    # Valider le modèle
    try:
        Model = apps.get_model('api', model_name)
    except LookupError:
        return HttpResponse(
            json.dumps({'error': f'Modèle {model_name} non trouvé'}),
            status=404,
            content_type='application/json'
        )
    
    # Récupérer les champs
    fields = get_model_fields(Model, exclude_auto=True)
    
    # Récupérer les données
    queryset = Model.objects.all()
    
    if format_type == 'xlsx':
        return _export_excel(model_name, fields, queryset)
    else:
        return _export_csv(model_name, fields, queryset)

def _export_csv(model_name, fields, queryset):
    """Exporte en CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{model_name}_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(fields)
    
    for obj in queryset:
        row = []
        for field_name in fields:
            value = getattr(obj, field_name, '')
            row.append(value)
        writer.writerow(row)
    
    return response

def _export_excel(model_name, fields, queryset):
    """Exporte en Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = model_name
    
    # En-tête
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col_num, field_name in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = field_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = 18
    
    # Données
    for row_num, obj in enumerate(queryset, start=2):
        for col_num, field_name in enumerate(fields, 1):
            value = getattr(obj, field_name, '')
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{model_name}_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response
